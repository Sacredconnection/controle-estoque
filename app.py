from __future__ import annotations

import csv
import io
import os
import secrets
import webbrowser
from functools import wraps
from pathlib import Path
from threading import Timer

from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

from qbo_stock.config import Settings
from qbo_stock.consolidation import consolidate_by_base_product, consolidate_inventory
from qbo_stock.db import Database
from qbo_stock.demo import DEMO_ITEMS
from qbo_stock.qbo import QBOError, QuickBooksService
from qbo_stock.security import TokenCipher, load_flask_secret

BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
INSTANCE_DIR.mkdir(exist_ok=True)

settings = Settings.from_env()
app = Flask(__name__, instance_path=str(INSTANCE_DIR), instance_relative_config=True)
app.secret_key = load_flask_secret(INSTANCE_DIR)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)
if settings.qbo_redirect_uri.lower().startswith("https://"):
    app.config["SESSION_COOKIE_SECURE"] = True

db = Database(INSTANCE_DIR / "quickbooks_stock.db")
cipher = TokenCipher(INSTANCE_DIR)
qbo = QuickBooksService(settings, db, cipher)


def protected(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if settings.app_password and not session.get("authenticated"):
            session["next_url"] = request.full_path if request.query_string else request.path
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


@app.template_filter("fmt3")
def fmt3(value):
    if value is None or value == "":
        return "—"
    try:
        return f"{float(value):.3f}"
    except (TypeError, ValueError):
        return str(value)


def _filtered_rows() -> tuple[list[dict], list[dict], dict]:
    rows, stats = consolidate_inventory(db.list_inventory(), settings.labels)
    q = request.args.get("q", "").strip().lower()
    hide_zero = request.args.get("hide_zero") == "1"
    only_unmatched = request.args.get("only_unmatched") == "1"
    only_weight_issues = request.args.get("only_weight_issues") == "1"

    if q:
        rows = [
            row
            for row in rows
            if q
            in " ".join(
                [
                    row["sku"],
                    row["product"],
                    row["sku_a"],
                    row["sku_b"],
                    row["weight_label"],
                    row["weight_source"],
                    row["status"],
                ]
            ).lower()
        ]
    if hide_zero:
        rows = [row for row in rows if abs(row["total"]) > 1e-12]
    if only_unmatched:
        rows = [row for row in rows if not row["status"].startswith("Nos dois")]
    if only_weight_issues:
        rows = [row for row in rows if row["unit_weight_kg"] is None]

    base_rows = consolidate_by_base_product(rows, settings.labels)
    return rows, base_rows, stats


@app.context_processor
def inject_globals():
    return {
        "labels": settings.labels,
        "qbo_environment": settings.qbo_environment,
        "credentials_ready": settings.credentials_ready,
        "legal_business_name": settings.legal_business_name,
        "legal_contact_email": settings.legal_contact_email,
        "legal_country": settings.legal_country,
    }


@app.route("/health")
def health():
    return {"status": "ok"}


@app.route("/launch")
def launch():
    return redirect(url_for("index"))


@app.route("/connect")
@app.route("/reconnect")
def connect_landing():
    return render_template("connect.html")


@app.route("/disconnect")
def disconnect_landing():
    return render_template("disconnect.html")


@app.route("/eula")
@app.route("/termos-de-uso")
def eula():
    return render_template("eula.html")


@app.route("/privacy")
@app.route("/politica-de-privacidade")
def privacy_policy():
    return render_template("privacy.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if not settings.app_password:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        supplied = request.form.get("password", "")
        if secrets.compare_digest(supplied, settings.app_password):
            session["authenticated"] = True
            target = session.pop("next_url", None) or url_for("index")
            return redirect(target)
        error = "Senha incorreta."
    return render_template("login.html", error=error)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@protected
def index():
    rows, base_rows, stats = _filtered_rows()
    connections = {slot: db.connection_summary(slot) for slot in ("A", "B")}
    any_demo = any(conn and conn.get("is_demo") for conn in connections.values())
    return render_template(
        "index.html",
        rows=rows,
        base_rows=base_rows,
        stats=stats,
        connections=connections,
        any_demo=any_demo,
        production_redirect_warning=settings.production_redirect_warning,
        search=request.args.get("q", ""),
        hide_zero=request.args.get("hide_zero") == "1",
        only_unmatched=request.args.get("only_unmatched") == "1",
        only_weight_issues=request.args.get("only_weight_issues") == "1",
    )


@app.route("/connect/<slot>")
@protected
def connect_company(slot: str):
    slot = slot.upper()
    if slot not in {"A", "B"}:
        flash("Empresa inválida.", "error")
        return redirect(url_for("index"))
    if not settings.credentials_ready:
        flash("Preencha as credenciais do Intuit no arquivo .env antes de conectar.", "error")
        return redirect(url_for("index"))

    state = qbo.new_state()
    session["oauth_state"] = state
    session["oauth_slot"] = slot
    try:
        return redirect(qbo.authorization_url(state))
    except QBOError as exc:
        flash(str(exc), "error")
        return redirect(url_for("index"))


@app.route("/oauth/callback")
def oauth_callback():
    error = request.args.get("error")
    if error:
        detail = request.args.get("error_description") or error
        flash(f"Autorização cancelada ou recusada: {detail}", "error")
        return redirect(url_for("index"))

    expected_state = session.pop("oauth_state", None)
    slot = session.pop("oauth_slot", None)
    received_state = request.args.get("state")
    code = request.args.get("code")
    realm_id = request.args.get("realmId")

    if not expected_state or not received_state or not secrets.compare_digest(
        expected_state, received_state
    ):
        flash("O retorno do OAuth não passou na validação de segurança (state).", "error")
        return redirect(url_for("index"))
    if slot not in {"A", "B"} or not code or not realm_id:
        flash("O QuickBooks não retornou code/realmId completos.", "error")
        return redirect(url_for("index"))

    try:
        token_data = qbo.exchange_code(code)
        qbo.save_authorization(slot, realm_id, token_data)
        count = qbo.sync_inventory(slot)
        flash(
            f"{settings.labels[slot]} conectada e sincronizada: {count} itens de estoque.",
            "success",
        )
    except Exception as exc:
        flash(f"Não foi possível concluir a conexão: {exc}", "error")
    return redirect(url_for("index"))


@app.route("/sync/<slot>", methods=["POST"])
@protected
def sync_company(slot: str):
    slot = slot.upper()
    if slot not in {"A", "B"}:
        flash("Empresa inválida.", "error")
        return redirect(url_for("index"))
    try:
        count = qbo.sync_inventory(slot)
        flash(f"{settings.labels[slot]} atualizada: {count} itens.", "success")
    except Exception as exc:
        flash(f"Erro ao atualizar {settings.labels[slot]}: {exc}", "error")
    return redirect(url_for("index"))


@app.route("/sync/all", methods=["POST"])
@protected
def sync_all():
    messages = []
    errors = []
    for slot in ("A", "B"):
        connection = db.get_connection(slot)
        if not connection:
            errors.append(f"{settings.labels[slot]} não está conectada")
            continue
        if connection.get("is_demo"):
            errors.append(f"{settings.labels[slot]} está em modo demonstração")
            continue
        try:
            count = qbo.sync_inventory(slot)
            messages.append(f"{settings.labels[slot]}: {count} itens")
        except Exception as exc:
            errors.append(f"{settings.labels[slot]}: {exc}")
    if messages:
        flash("Atualização concluída — " + "; ".join(messages), "success")
    if errors:
        flash("Não atualizado — " + "; ".join(errors), "error")
    return redirect(url_for("index"))


@app.route("/disconnect/<slot>", methods=["POST"])
@protected
def disconnect_company(slot: str):
    slot = slot.upper()
    if slot in {"A", "B"}:
        db.delete_connection(slot)
        flash(f"{settings.labels[slot]} removida deste programa.", "success")
    return redirect(url_for("index"))


@app.route("/demo/load", methods=["POST"])
@protected
def demo_load():
    db.load_demo(settings.labels, DEMO_ITEMS)
    flash("Dados de demonstração com potes de 250g e 500g carregados.", "success")
    return redirect(url_for("index"))


@app.route("/demo/clear", methods=["POST"])
@protected
def demo_clear():
    db.clear_demo()
    flash("Dados de demonstração removidos.", "success")
    return redirect(url_for("index"))


@app.route("/export.csv")
@protected
def export_csv():
    rows, _ = consolidate_inventory(db.list_inventory(), settings.labels)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(
        [
            "SKU consolidado",
            "Produto",
            "Peso por unidade",
            "Peso unitário (kg)",
            "Origem do peso",
            f"SKU {settings.labels['A']}",
            f"Unidades {settings.labels['A']}",
            f"Kg {settings.labels['A']}",
            f"SKU {settings.labels['B']}",
            f"Unidades {settings.labels['B']}",
            f"Kg {settings.labels['B']}",
            "Unidades totais",
            "Kg total",
            "Conferência",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row["sku"],
                row["product"],
                row["weight_label"],
                "" if row["unit_weight_kg"] is None else row["unit_weight_kg"],
                row["weight_source"],
                row["sku_a"],
                row["qty_a"],
                "" if row["kg_a"] is None else row["kg_a"],
                row["sku_b"],
                row["qty_b"],
                "" if row["kg_b"] is None else row["kg_b"],
                row["total"],
                "" if row["kg_total"] is None else row["kg_total"],
                row["status"],
            ]
        )
    data = "\ufeff" + output.getvalue()
    return send_file(
        io.BytesIO(data.encode("utf-8")),
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name="estoque_quickbooks_unidades_e_kg.csv",
    )


def _style_header(row):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


@app.route("/export.xlsx")
@protected
def export_xlsx():
    rows, stats = consolidate_inventory(db.list_inventory(), settings.labels)
    base_rows = consolidate_by_base_product(rows, settings.labels)
    book = Workbook()
    sheet = book.active
    sheet.title = "Estoque por SKU"

    title = "Estoque consolidado — unidades e conversão para kg"
    sheet.merge_cells("A1:N1")
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "SKU consolidado",
        "Produto",
        "Peso por unidade",
        "Peso unitário (kg)",
        "Origem do peso",
        f"SKU {settings.labels['A']}",
        f"Unidades {settings.labels['A']}",
        f"Kg {settings.labels['A']}",
        f"SKU {settings.labels['B']}",
        f"Unidades {settings.labels['B']}",
        f"Kg {settings.labels['B']}",
        "Unidades totais",
        "Kg total",
        "Conferência",
    ]
    sheet.append([])
    sheet.append(headers)
    _style_header(sheet[3])

    for row in rows:
        sheet.append(
            [
                row["sku"],
                row["product"],
                row["weight_label"],
                row["unit_weight_kg"],
                row["weight_source"],
                row["sku_a"],
                row["qty_a"],
                row["kg_a"],
                row["sku_b"],
                row["qty_b"],
                row["kg_b"],
                row["total"],
                row["kg_total"],
                row["status"],
            ]
        )

    for row_number in range(4, sheet.max_row + 1):
        for col in (4, 7, 8, 10, 11, 12, 13):
            sheet.cell(row_number, col).number_format = "0.000"

    sheet.auto_filter.ref = f"A3:N{sheet.max_row}"
    sheet.freeze_panes = "A4"
    widths = [18, 42, 16, 18, 38, 20, 17, 15, 20, 17, 15, 17, 15, 48]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    base = book.create_sheet("Resumo por Rapé")
    base.merge_cells("A1:I1")
    base["A1"] = "Resumo em kg por produto-base"
    base["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    base["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    base["A1"].alignment = Alignment(horizontal="center")
    base.append([])
    base.append(
        [
            "Código-base",
            "Produto-base",
            f"Potes/unidades {settings.labels['A']}",
            f"Kg {settings.labels['A']}",
            f"Potes/unidades {settings.labels['B']}",
            f"Kg {settings.labels['B']}",
            "Kg total",
            "Variações incluídas",
            "Conferência",
        ]
    )
    _style_header(base[3])
    for row in base_rows:
        base.append(
            [
                row["base_sku"],
                row["product"],
                row["qty_a"],
                row["kg_a"],
                row["qty_b"],
                row["kg_b"],
                row["kg_total"],
                row["variants"],
                row["status"],
            ]
        )
    for row_number in range(4, base.max_row + 1):
        for col in (3, 4, 5, 6, 7):
            base.cell(row_number, col).number_format = "0.000"
    base.auto_filter.ref = f"A3:I{base.max_row}"
    base.freeze_panes = "A4"
    base_widths = [18, 42, 20, 16, 20, 16, 16, 58, 38]
    for index, width in enumerate(base_widths, start=1):
        base.column_dimensions[get_column_letter(index)].width = width

    summary = book.create_sheet("Resumo")
    summary.append(["Indicador", "Valor"])
    summary.append(["SKUs consolidados", stats["products"]])
    summary.append([f"QtyOnHand somado — {settings.labels['A']}", stats["qty_a"]])
    summary.append([f"QtyOnHand somado — {settings.labels['B']}", stats["qty_b"]])
    summary.append(["QtyOnHand somado — total", stats["qty_total"]])
    summary.append([f"Kg identificados — {settings.labels['A']}", stats["kg_a"]])
    summary.append([f"Kg identificados — {settings.labels['B']}", stats["kg_b"]])
    summary.append(["Kg identificados — total", stats["kg_total"]])
    summary.append(["Produtos com peso identificado", stats["weight_known"]])
    summary.append(["Produtos sem peso identificado", stats["weight_unknown"]])
    summary.append(["Produtos com conflito de peso", stats["weight_conflicts"]])
    summary.append(["Presentes nos dois", stats["both"]])
    summary.append([f"Somente {settings.labels['A']}", stats["only_a"]])
    summary.append([f"Somente {settings.labels['B']}", stats["only_b"]])
    summary.append(["Sem SKU", stats["no_sku"]])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 20
    for row_number in range(2, summary.max_row + 1):
        summary.cell(row_number, 2).number_format = "0.000"

    stream = io.BytesIO()
    book.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="estoque_quickbooks_unidades_e_kg.xlsx",
    )


def _open_browser() -> None:
    webbrowser.open_new(f"http://localhost:{settings.port}")


if __name__ == "__main__":
    if os.getenv("OPEN_BROWSER", "true").strip().lower() in {"1", "true", "yes", "sim"}:
        Timer(1.2, _open_browser).start()
    app.run(host=settings.host, port=settings.port, debug=settings.debug)
